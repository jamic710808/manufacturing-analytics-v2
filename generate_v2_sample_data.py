#!/usr/bin/env python3
"""
製造業分析儀表板 V2 — 範例資料生成器
產出：Manufacturing_Sample_Data_V2.xlsx（7 個工作表）
"""
import random, datetime, math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)

wb = Workbook()

# ===== 共用常數 =====
SUPPLIERS = [
    {"name":"鑫達材料","category":"金屬","tier":"Preferred","qualityScore":92,"serviceScore":88,"techScore":85},
    {"name":"宏遠化工","category":"化學品","tier":"Preferred","qualityScore":90,"serviceScore":91,"techScore":87},
    {"name":"永豐塑膠","category":"塑膠","tier":"Approved","qualityScore":85,"serviceScore":82,"techScore":78},
    {"name":"正大五金","category":"五金","tier":"Approved","qualityScore":83,"serviceScore":80,"techScore":75},
    {"name":"聯合電子","category":"電子","tier":"Preferred","qualityScore":95,"serviceScore":93,"techScore":92},
    {"name":"明輝包裝","category":"包裝","tier":"Approved","qualityScore":80,"serviceScore":78,"techScore":70},
    {"name":"瑞豐鋼鐵","category":"金屬","tier":"At-Risk","qualityScore":72,"serviceScore":65,"techScore":68},
    {"name":"東昇橡膠","category":"橡膠","tier":"Approved","qualityScore":82,"serviceScore":79,"techScore":76},
    {"name":"國際化學","category":"化學品","tier":"At-Risk","qualityScore":70,"serviceScore":60,"techScore":65},
    {"name":"精密模具","category":"模具","tier":"Preferred","qualityScore":94,"serviceScore":90,"techScore":95},
    {"name":"大同塗料","category":"塗料","tier":"Approved","qualityScore":81,"serviceScore":77,"techScore":73},
    {"name":"富邦物流","category":"物流","tier":"Approved","qualityScore":78,"serviceScore":85,"techScore":72},
]

MATERIALS = [
    {"id":"M001","name":"不鏽鋼板 SUS304","unitCost":185,"abcClass":"A","type":"通用料","unit":"KG","safetyStock":500,"reorderPoint":800,"leadTimeDays":14},
    {"id":"M002","name":"銅線 C1100","unitCost":320,"abcClass":"A","type":"通用料","unit":"KG","safetyStock":300,"reorderPoint":500,"leadTimeDays":10},
    {"id":"M003","name":"ABS 塑膠粒","unitCost":65,"abcClass":"B","type":"通用料","unit":"KG","safetyStock":1000,"reorderPoint":1500,"leadTimeDays":7},
    {"id":"M004","name":"環氧樹脂 EP-200","unitCost":420,"abcClass":"A","type":"專用料","unit":"KG","safetyStock":200,"reorderPoint":350,"leadTimeDays":21},
    {"id":"M005","name":"鋁合金 6061","unitCost":145,"abcClass":"B","type":"通用料","unit":"KG","safetyStock":600,"reorderPoint":900,"leadTimeDays":12},
    {"id":"M006","name":"矽膠墊片","unitCost":28,"abcClass":"C","type":"專用料","unit":"PCS","safetyStock":2000,"reorderPoint":3000,"leadTimeDays":5},
    {"id":"M007","name":"電容 100μF","unitCost":3.5,"abcClass":"C","type":"通用料","unit":"PCS","safetyStock":5000,"reorderPoint":8000,"leadTimeDays":7},
    {"id":"M008","name":"PCB 基板 FR-4","unitCost":85,"abcClass":"B","type":"專用料","unit":"PCS","safetyStock":800,"reorderPoint":1200,"leadTimeDays":14},
    {"id":"M009","name":"螺絲 M3×8","unitCost":0.8,"abcClass":"C","type":"通用料","unit":"PCS","safetyStock":10000,"reorderPoint":15000,"leadTimeDays":3},
    {"id":"M010","name":"防鏽漆 RAL7035","unitCost":210,"abcClass":"B","type":"通用料","unit":"L","safetyStock":100,"reorderPoint":200,"leadTimeDays":10},
    {"id":"M011","name":"橡膠O型環","unitCost":5.2,"abcClass":"C","type":"專用料","unit":"PCS","safetyStock":3000,"reorderPoint":5000,"leadTimeDays":5},
    {"id":"M012","name":"鍍鋅鋼管","unitCost":95,"abcClass":"B","type":"通用料","unit":"M","safetyStock":200,"reorderPoint":400,"leadTimeDays":14},
    {"id":"M013","name":"導熱膏 TC-500","unitCost":580,"abcClass":"A","type":"專用料","unit":"KG","safetyStock":50,"reorderPoint":100,"leadTimeDays":28},
    {"id":"M014","name":"包裝紙箱 A3","unitCost":12,"abcClass":"C","type":"非生產用料","unit":"PCS","safetyStock":500,"reorderPoint":1000,"leadTimeDays":3},
    {"id":"M015","name":"標籤貼紙","unitCost":1.5,"abcClass":"C","type":"非生產用料","unit":"PCS","safetyStock":2000,"reorderPoint":5000,"leadTimeDays":3},
]

PRODUCTS = [
    {"id":"P001","name":"精密控制器 CTL-A","bom":[("M001",2.5),("M004",0.8),("M007",12),("M008",1),("M009",8)]},
    {"id":"P002","name":"電源模組 PWR-B","bom":[("M002",1.2),("M003",0.5),("M007",20),("M008",2),("M006",4)]},
    {"id":"P003","name":"散熱組件 HSK-C","bom":[("M005",3.0),("M013",0.3),("M011",6),("M009",12),("M006",2)]},
]

LINES = ["A線","B線","C線"]
COST_TYPES = [
    {"level1":"直接材料","level2":"金屬材料","level3":"不鏽鋼"},
    {"level1":"直接材料","level2":"金屬材料","level3":"銅材"},
    {"level1":"直接材料","level2":"金屬材料","level3":"鋁材"},
    {"level1":"直接材料","level2":"化學品","level3":"樹脂"},
    {"level1":"直接材料","level2":"電子零件","level3":"被動元件"},
    {"level1":"直接材料","level2":"電子零件","level3":"PCB"},
    {"level1":"直接人工","level2":"生產人工","level3":"組裝工資"},
    {"level1":"直接人工","level2":"生產人工","level3":"測試工資"},
    {"level1":"直接人工","level2":"品管人工","level3":"檢驗工資"},
    {"level1":"製造費用","level2":"設備折舊","level3":"產線設備"},
    {"level1":"製造費用","level2":"能源費用","level3":"電力"},
    {"level1":"製造費用","level2":"間接材料","level3":"耗材"},
]

# ===== 樣式工具 =====
header_font = Font(name='Microsoft JhengHei', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)

def style_header(ws, ncols):
    for col in range(1, ncols+1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

# ===== Sheet 1: Materials =====
ws = wb.active
ws.title = 'Materials'
mat_headers = ['id','name','unitCost','stockQty','abcClass','daysInStock','type','monthlyUsage','unit','safetyStock','reorderPoint','leadTimeDays','lastMovementDate']
ws.append(mat_headers)
base_date = datetime.date(2025, 12, 31)
for m in MATERIALS:
    stock_qty = random.randint(100, 5000)
    days_in_stock = random.randint(5, 180)
    monthly_usage = random.randint(50, 2000)
    last_move = base_date - datetime.timedelta(days=random.randint(1, 90))
    ws.append([m['id'], m['name'], m['unitCost'], stock_qty, m['abcClass'], days_in_stock,
               m['type'], monthly_usage, m['unit'], m['safetyStock'], m['reorderPoint'],
               m['leadTimeDays'], last_move.isoformat()])
style_header(ws, len(mat_headers))
auto_width(ws)

# ===== Sheet 2: Suppliers =====
ws2 = wb.create_sheet('Suppliers')
sup_headers = ['name','category','rating','onTimeDeliveryRate','tier','qualityScore','serviceScore','techScore']
ws2.append(sup_headers)
for s in SUPPLIERS:
    rating = round(random.uniform(3.0, 5.0), 1)
    otd = round(random.uniform(0.65, 0.99), 2)
    ws2.append([s['name'], s['category'], rating, otd, s['tier'], s['qualityScore'], s['serviceScore'], s['techScore']])
style_header(ws2, len(sup_headers))
auto_width(ws2)

# ===== Sheet 3: Products =====
ws3 = wb.create_sheet('Products')
prod_headers = ['productId','productName','bomMaterialId','bomMaterialName','bomQtyPerUnit']
ws3.append(prod_headers)
mat_dict = {m['id']: m['name'] for m in MATERIALS}
for p in PRODUCTS:
    for mid, qty in p['bom']:
        ws3.append([p['id'], p['name'], mid, mat_dict.get(mid, mid), qty])
style_header(ws3, len(prod_headers))
auto_width(ws3)

# ===== Sheet 4: PurchaseOrders (300 筆) =====
ws4 = wb.create_sheet('PurchaseOrders')
po_headers = ['poId','supplierName','materialId','materialName','qty','pricePerUnit','totalAmount',
              'orderDate','plannedDeliveryDate','actualDeliveryDate','delayDays']
ws4.append(po_headers)
for i in range(1, 301):
    sup = random.choice(SUPPLIERS)
    mat = random.choice(MATERIALS)
    qty = random.randint(50, 3000)
    price = round(mat['unitCost'] * random.uniform(0.85, 1.25), 2)
    total = round(qty * price, 2)
    order_date = datetime.date(2025, 1, 1) + datetime.timedelta(days=random.randint(0, 364))
    planned = order_date + datetime.timedelta(days=mat['leadTimeDays'])
    delay = random.choices([0, 0, 0, random.randint(1, 15), random.randint(-5, -1)], weights=[50, 20, 10, 15, 5])[0]
    actual = planned + datetime.timedelta(days=delay)
    ws4.append([f'PO-{i:04d}', sup['name'], mat['id'], mat['name'], qty, price, total,
                order_date.isoformat(), planned.isoformat(), actual.isoformat(), delay])
style_header(ws4, len(po_headers))
auto_width(ws4)

# ===== Sheet 5: ProductionOrders (150 筆) =====
ws5 = wb.create_sheet('ProductionOrders')
prod_headers2 = ['orderId','productId','productName','line','orderDate','inputQty','goodOutput',
                 'defectQty','scrapQty','reworkQty','yieldRate',
                 'stdMaterialCost','actMaterialCost','materialVar',
                 'stdLaborCost','actLaborCost','laborVar',
                 'stdOverheadCost','actOverheadCost','overheadVar','actTotalCost']
ws5.append(prod_headers2)
for i in range(1, 151):
    prod = random.choice(PRODUCTS)
    line = random.choice(LINES)
    order_date = datetime.date(2025, 1, 1) + datetime.timedelta(days=random.randint(0, 364))
    input_qty = random.randint(100, 1000)
    yield_rate = round(random.uniform(0.88, 0.995), 3)
    good = int(input_qty * yield_rate)
    defect = random.randint(0, input_qty - good)
    scrap = input_qty - good - defect
    rework = random.randint(0, defect)
    std_mat = round(input_qty * random.uniform(80, 250), 2)
    act_mat = round(std_mat * random.uniform(0.92, 1.12), 2)
    mat_var = round(act_mat - std_mat, 2)
    std_lab = round(input_qty * random.uniform(15, 45), 2)
    act_lab = round(std_lab * random.uniform(0.95, 1.08), 2)
    lab_var = round(act_lab - std_lab, 2)
    std_oh = round(input_qty * random.uniform(20, 60), 2)
    act_oh = round(std_oh * random.uniform(0.93, 1.10), 2)
    oh_var = round(act_oh - std_oh, 2)
    act_total = round(act_mat + act_lab + act_oh, 2)
    ws5.append([f'WO-{i:04d}', prod['id'], prod['name'], line, order_date.isoformat(),
                input_qty, good, defect, scrap, rework, yield_rate,
                std_mat, act_mat, mat_var, std_lab, act_lab, lab_var,
                std_oh, act_oh, oh_var, act_total])
style_header(ws5, len(prod_headers2))
auto_width(ws5)

# ===== Sheet 6: InventorySnapshots (12 個月 × 15 物料 = 180 筆) =====
ws6 = wb.create_sheet('InventorySnapshots')
snap_headers = ['month','materialId','materialName','openingQty','inboundQty','outboundQty','closingQty','closingValue']
ws6.append(snap_headers)
for m_idx in range(1, 13):
    month_str = f'2025-{m_idx:02d}'
    for mat in MATERIALS:
        opening = random.randint(200, 3000)
        inbound = random.randint(100, 2000)
        outbound = random.randint(100, min(opening + inbound, 2500))
        closing = opening + inbound - outbound
        closing_val = round(closing * mat['unitCost'], 2)
        ws6.append([month_str, mat['id'], mat['name'], opening, inbound, outbound, closing, closing_val])
style_header(ws6, len(snap_headers))
auto_width(ws6)

# ===== Sheet 7: CostTypes =====
ws7 = wb.create_sheet('CostTypes')
ct_headers = ['level1','level2','level3']
ws7.append(ct_headers)
for ct in COST_TYPES:
    ws7.append([ct['level1'], ct['level2'], ct['level3']])
style_header(ws7, len(ct_headers))
auto_width(ws7)

# ===== 儲存 =====
output_path = r'c:\Users\jamic\製造業分析\Manufacturing_Sample_Data_V2.xlsx'
wb.save(output_path)
print(f'[OK] saved to {output_path}')
print(f'  sheets: {wb.sheetnames}')
print(f'  Materials: {ws.max_row-1}')
print(f'  Suppliers: {ws2.max_row-1}')
print(f'  Products (BOM): {ws3.max_row-1}')
print(f'  PurchaseOrders: {ws4.max_row-1}')
print(f'  ProductionOrders: {ws5.max_row-1}')
print(f'  InventorySnapshots: {ws6.max_row-1}')
print(f'  CostTypes: {ws7.max_row-1}')
